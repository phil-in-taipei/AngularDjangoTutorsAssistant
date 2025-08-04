import io
from openpyxl import Workbook
from openpyxl.styles import Alignment
from django.core.mail import EmailMessage


def send_class_data_excel_via_email(
        data, month, year, email_address, 
        teacher_name, school_name
    ):
    # Create workbook in memory
    wb = Workbook()
    ws = wb.active
    ws.title = "10"
    date_string = F'{month}/{year}'
    file_name = F'{teacher_name}_{school_name}_{date_string}.xlsx'

    # write date, teacher's name, and email address
    ws.cell(1, column=1, value=date_string)
    ws.cell(1, column=2, value=teacher_name)
    ws.cell(1, column=9, value=email_address)

    # Set column B width
    ws.column_dimensions['B'].width = 32.5

    # Write headers
    headers = [
        "", "Class", "HR per times(H)", "Times(T)",
        "Total Hrs(TH=T*H)", "Pay per HR(P)", "Taxi(A)", "Total Pay(TP=TH*P+A*T)"
    ]
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_num, value=header)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    total_pay = 0
    total_hours = 0
    last_row = 0

    # Write data rows
    for row_num, entry in enumerate(data, start=3):
        ws.cell(row=row_num, column=2, value=f"{entry['student_or_class_name']} {entry['scheduled_classes']}")
        ws.cell(row=row_num, column=3, value=entry['class_duration'])
        ws.cell(row=row_num, column=4, value=entry['number_of_classes'])
        ws.cell(row=row_num, column=5, value=entry['total_hours'])
        ws.cell(row=row_num, column=6, value=entry['pay rate'])
        ws.cell(row=row_num, column=7, value=0)  # Taxi(A)
        ws.cell(row=row_num, column=8, value=entry['payment'])
        total_pay += entry['payment']
        total_hours += entry['total_hours']
        last_row = row_num

    ws.cell(row=last_row + 1, column=1, value="FS")
    ws.cell(row=last_row + 1, column=5, value=total_hours)
    ws.cell(row=last_row + 1, column=8, value=total_pay)


    # Save to BytesIO buffer
    excel_stream = io.BytesIO()
    wb.save(excel_stream)
    excel_stream.seek(0)

    # Send email using Django's default from_email
    try:
        email = EmailMessage(
            subject=f'{school_name} Monthly Report: {month}/{year}',
            body=f"""Dear {teacher_name},

        Please find the attached file with your monthly hours payment 
        information at {school_name} in {date_string}

        Best,
        -Teacher's Assistant""",
            to=[email_address]
        )
        email.attach(
            file_name, excel_stream.read(),
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        email.send()
    except Exception as e:
        print("Email send error:", e)
